import io
import re
from collections import defaultdict
from datetime import date, datetime
from typing import List

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from openpyxl.utils.datetime import from_excel
from sqlalchemy import extract, func
from sqlalchemy.orm import Session

from auth import get_current_active_user
from database import get_db
from models import FinanceRecord, FleetOrganization, Mine
from schemas.finance import FinanceCreate, FinanceOut, FinanceUpdate
from utils.crud_helpers import get_object_or_404
from utils.excel_utils import create_export, create_template
from utils.permissions import check_mine_access, filter_by_mine, get_target_mine_id
from utils.fleet_scope import resolve_single_fleet_id
from services.data_governance import archive_record, assert_period_unlocked

router = APIRouter()

FINANCE_FIELDS = ["trans_type", "amount", "currency", "category", "description", "recorder", "trans_date"]
FINANCE_HEADERS = ["类型(收入/支出)", "金额", "货币(USD/CDF)", "类别", "描述", "记录人", "日期"]
FINANCE_COLUMN_WIDTHS = [18, 14, 16, 16, 24, 16, 14]
VALID_CURRENCIES = {"USD", "CDF"}
EXPENSE_CATEGORIES = ("外联", "工资", "其他")
EXPENSE_CATEGORY_ALIASES = {
    "外联": "外联",
    "外联费": "外联",
    "外协": "外联",
    "公关": "外联",
    "招待": "外联",
    "关系": "外联",
    "工资": "工资",
    "薪资": "工资",
    "薪水": "工资",
    "人工": "工资",
    "劳务": "工资",
    "其他": "其他",
    "其它": "其他",
}

HEADER_ALIASES = {
    "trans_type": ["类型(收入/支出)", "类型", "收支类型", "收入/支出", "收入支出", "支出收入", "交易类型", "trans_type", "type"],
    "amount": ["金额", "金額", "数额", "费用", "收入金额", "支出金额", "amount", "money"],
    "currency": ["货币(USD/CDF)", "货币", "币种", "币别", "currency"],
    "category": ["类别", "分类", "支出类别", "收入类别", "模块", "category"],
    "description": ["描述", "说明", "备注", "用途", "摘要", "description", "remark"],
    "recorder": ["记录人", "录入人", "经办人", "recorder", "created_by"],
    "trans_date": ["日期", "交易日期", "发生日期", "记录日期", "trans_date", "date"],
}


def _get_and_check(db, model, obj_id, current_user):
    obj = get_object_or_404(db, model, obj_id)
    _check_scope(current_user, obj)
    return obj


def _first_mine_id(db: Session) -> str | None:
    first_mine = db.query(Mine).order_by(Mine.created_at).first()
    return first_mine.id if first_mine else None


def _get_fleet(db: Session, fleet_id: str | None) -> FleetOrganization | None:
    if not fleet_id:
        return None
    fleet = db.query(FleetOrganization).filter(FleetOrganization.id == fleet_id, FleetOrganization.active == 1).first()
    if not fleet:
        raise HTTPException(status_code=400, detail="Fleet organization does not exist or is disabled")
    return fleet


def _scope_for_request(db: Session, current_user, mine_id: str | None = None, fleet_id: str | None = None) -> dict:
    fleet_id = resolve_single_fleet_id(db, fleet_id)
    if current_user.role == "fleet":
        if not current_user.fleet_id:
            raise HTTPException(status_code=403, detail="Fleet account is not bound to a fleet")
        fleet = _get_fleet(db, current_user.fleet_id)
        target_mine = fleet.mine_id or _first_mine_id(db)
        if not target_mine:
            raise HTTPException(status_code=400, detail="No backing mine exists for fleet data")
        return {"mine_id": target_mine, "fleet_id": fleet.id}
    if fleet_id:
        if current_user.role != "super":
            raise HTTPException(status_code=403, detail="Only super admin can select fleet scope")
        fleet = _get_fleet(db, fleet_id)
        target_mine = fleet.mine_id or _first_mine_id(db)
        if not target_mine:
            raise HTTPException(status_code=400, detail="No backing mine exists for fleet data")
        return {"mine_id": target_mine, "fleet_id": fleet.id}
    target_mine = get_target_mine_id(current_user, mine_id)
    if target_mine is None:
        raise HTTPException(status_code=400, detail="mine_id is required")
    check_mine_access(current_user, target_mine)
    return {"mine_id": target_mine, "fleet_id": None}


def _filter_scope(query, current_user, mine_id: str | None = None, fleet_id: str | None = None):
    fleet_id = resolve_single_fleet_id(query.session, fleet_id)
    if current_user.role == "fleet":
        return query.filter(FinanceRecord.fleet_id == current_user.fleet_id)
    if fleet_id:
        if current_user.role != "super":
            raise HTTPException(status_code=403, detail="Only super admin can select fleet scope")
        return query.filter(FinanceRecord.fleet_id == fleet_id)
    query = filter_by_mine(query, FinanceRecord, "mine_id", current_user, mine_id)
    return query.filter(FinanceRecord.fleet_id == None)


def _check_scope(current_user, obj):
    if current_user.role == "fleet":
        if obj.fleet_id != current_user.fleet_id:
            raise HTTPException(status_code=403, detail="Permission denied")
        return
    if obj.fleet_id and current_user.role != "super":
        raise HTTPException(status_code=403, detail="Permission denied")
    if not obj.fleet_id:
        check_mine_access(current_user, obj.mine_id)


def _enum_value(value):
    return value.value if hasattr(value, "value") else value


def _normalize_header(value):
    return re.sub(r"[\s()（）/\\:：_\-]+", "", str(value or "").strip().lower())


HEADER_TO_FIELD = {
    _normalize_header(alias): field
    for field, aliases in HEADER_ALIASES.items()
    for alias in aliases
}


def _build_column_map(ws):
    column_map = {}
    for column in range(1, ws.max_column + 1):
        field = HEADER_TO_FIELD.get(_normalize_header(ws.cell(row=1, column=column).value))
        if field and field not in column_map:
            column_map[field] = column

    missing = [field for field in ("trans_type", "amount", "trans_date") if field not in column_map]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=(
                "Excel表头不匹配，请使用模板或包含这些列：类型、金额、日期。"
                f" 当前识别到: {[ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]}"
            )
        )
    return column_map


def _cell_value(ws, row, column_map, field, default=""):
    column = column_map.get(field)
    if not column:
        return default
    value = ws.cell(row=row, column=column).value
    return default if value is None else value


def _is_placeholder_or_blank_row(ws, row, column_map):
    values = [_cell_value(ws, row, column_map, field, "") for field in column_map]
    if all(value in (None, "") for value in values):
        return True
    return all(str(value).strip() == "在此填写数据" for value in values if value not in (None, ""))


def _normalize_currency(value):
    text = str(value or "USD").strip().upper()
    aliases = {
        "$": "USD",
        "US$": "USD",
        "美元": "USD",
        "美金": "USD",
        "FC": "CDF",
        "FCFA": "CDF",
        "刚郎": "CDF",
        "刚果法郎": "CDF",
    }
    currency = aliases.get(text, text)
    return currency if currency in VALID_CURRENCIES else None


def _normalize_trans_type(value):
    text = str(value or "").strip().lower()
    compact = re.sub(r"\s+", "", text)
    if compact in {"income", "in", "收入", "收", "进账", "入账"} or "收入" in compact:
        return "income"
    if compact in {"expense", "out", "支出", "支", "付款", "开支"} or "支出" in compact:
        return "expense"
    return None


def _normalize_expense_category(value):
    text = str(value or "").strip()
    compact = re.sub(r"\s+", "", text)
    if not compact:
        return None
    for alias, category in EXPENSE_CATEGORY_ALIASES.items():
        if alias == compact or alias in compact:
            return category
    return None


def _normalize_category(trans_type, value):
    category = str(value or "").strip()
    if trans_type != "expense":
        return category

    normalized = _normalize_expense_category(category)
    if normalized:
        return normalized
    raise HTTPException(status_code=400, detail="支出类别必须为：外联、工资或其他")


def _parse_amount(value):
    if value is None or value == "":
        raise ValueError("金额为空")
    if isinstance(value, str):
        text = value.strip().upper().replace(",", "").replace("，", "").replace("$", "").replace("USD", "").replace("CDF", "")
        text = re.sub(r"\s+", "", text)
        if text.startswith("(") and text.endswith(")"):
            text = "-" + text[1:-1]
        value = text
    amount = float(value)
    if amount == 0:
        raise ValueError("金额不能为0")
    return abs(amount)


def _parse_trans_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return from_excel(value).date()

    text = str(value or "").strip()
    if not text:
        raise ValueError("日期为空")
    normalized = text.replace("年", "-").replace("月", "-").replace("日", "").replace("/", "-").replace(".", "-")
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d-%m-%Y", "%d-%m-%y"):
        try:
            return datetime.strptime(normalized, fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(normalized).date()
    except ValueError as exc:
        raise ValueError("日期格式错误，请使用 YYYY-MM-DD") from exc


async def _load_finance_sheet(file: UploadFile):
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 格式的Excel文件")
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Excel文件读取失败: {exc}") from exc
    return wb.active


def _display_trans_type(value):
    if value == "income":
        return "收入"
    return "支出"


# ========== Finance CRUD (List/Create) ==========

@router.get("/", response_model=List[FinanceOut])
def read_finances(
    skip: int = 0,
    limit: int = 100,
    mine_id: str = None,
    fleet_id: str = None,
    trans_type: str = None,
    start_date: str = None,
    end_date: str = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user)
):
    query = _filter_scope(db.query(FinanceRecord), current_user, mine_id, fleet_id)
    if trans_type:
        query = query.filter(FinanceRecord.trans_type == trans_type)
    start_day = _parse_query_date(start_date, "开始日期")
    end_day = _parse_query_date(end_date, "结束日期")
    if start_day:
        query = query.filter(FinanceRecord.trans_date >= start_day)
    if end_day:
        query = query.filter(FinanceRecord.trans_date <= end_day)
    return query.order_by(FinanceRecord.trans_date.desc()).offset(skip).limit(limit).all()


@router.post("/", response_model=FinanceOut)
def create_finance(
    finance: FinanceCreate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user)
):
    scope = _scope_for_request(db, current_user, finance.mine_id, finance.fleet_id)
    target_mine = scope["mine_id"]
    target_fleet = scope["fleet_id"]
    assert_period_unlocked(db, target_mine, "finance", finance.trans_date)

    trans_type = _enum_value(finance.trans_type)
    category = _normalize_category(trans_type, finance.category)
    db_finance = FinanceRecord(
        mine_id=target_mine,
        fleet_id=target_fleet,
        trans_type=trans_type,
        amount=finance.amount,
        currency=_enum_value(finance.currency),
        category=category,
        description=finance.description or "",
        recorder=finance.recorder or "",
        trans_date=finance.trans_date
    )
    db.add(db_finance)
    db.commit()
    db.refresh(db_finance)
    return db_finance


# ========== Excel Import/Export ==========

@router.get("/expense-categories")
def get_finance_expense_categories(current_user=Depends(get_current_active_user)):
    return {"categories": list(EXPENSE_CATEGORIES)}


@router.get("/import/template")
def download_finance_template(current_user=Depends(get_current_active_user)):
    return create_template(FINANCE_HEADERS, "财务导入模板.xlsx", FINANCE_COLUMN_WIDTHS)


@router.post("/import/excel")
async def import_finance_excel(
    mine_id: str = None,
    fleet_id: str = None,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user)
):
    scope = _scope_for_request(db, current_user, mine_id, fleet_id)
    target_mine = scope["mine_id"]
    target_fleet = scope["fleet_id"]

    ws = await _load_finance_sheet(file)
    column_map = _build_column_map(ws)
    imported = 0
    errors = []

    for row in range(2, ws.max_row + 1):
        if _is_placeholder_or_blank_row(ws, row, column_map):
            continue

        raw_trans_type = _cell_value(ws, row, column_map, "trans_type")
        raw_amount = _cell_value(ws, row, column_map, "amount")
        raw_currency = _cell_value(ws, row, column_map, "currency", "USD")
        raw_category = _cell_value(ws, row, column_map, "category", "")
        description = _cell_value(ws, row, column_map, "description", "")
        recorder = _cell_value(ws, row, column_map, "recorder", "")
        raw_trans_date = _cell_value(ws, row, column_map, "trans_date")

        trans_type = _normalize_trans_type(raw_trans_type)
        if not trans_type:
            errors.append(f"第{row}行: 类型必须为收入或支出，当前值: {raw_trans_type}")
            continue

        currency = _normalize_currency(raw_currency)
        if not currency:
            errors.append(f"第{row}行: 货币必须为 USD 或 CDF，当前值: {raw_currency}")
            continue

        try:
            amount = _parse_amount(raw_amount)
            trans_date = _parse_trans_date(raw_trans_date)
            category = _normalize_category(trans_type, raw_category)
            assert_period_unlocked(db, target_mine, "finance", trans_date)
        except (TypeError, ValueError, HTTPException) as exc:
            detail = exc.detail if isinstance(exc, HTTPException) else str(exc)
            errors.append(f"第{row}行: {detail}")
            continue

        db.add(FinanceRecord(
            mine_id=target_mine,
            fleet_id=target_fleet,
            trans_type=trans_type,
            amount=amount,
            currency=currency,
            category=category,
            description=str(description or ""),
            recorder=str(recorder or ""),
            trans_date=trans_date
        ))
        imported += 1

    if imported:
        db.commit()
    return {"message": f"成功导入 {imported} 条记录", "imported": imported, "errors": errors}


@router.get("/export/excel")
def export_finance_excel(
    mine_id: str = None,
    fleet_id: str = None,
    trans_type: str = None,
    start_date: str = None,
    end_date: str = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user)
):
    query = _filter_scope(db.query(FinanceRecord), current_user, mine_id, fleet_id)
    if trans_type:
        query = query.filter(FinanceRecord.trans_type == trans_type)
    start_day = _parse_query_date(start_date, "开始日期")
    end_day = _parse_query_date(end_date, "结束日期")
    if start_day:
        query = query.filter(FinanceRecord.trans_date >= start_day)
    if end_day:
        query = query.filter(FinanceRecord.trans_date <= end_day)
    records = query.order_by(FinanceRecord.trans_date.desc()).all()
    rows = [
        [_display_trans_type(r.trans_type), r.amount, r.currency, r.category, r.description, r.recorder, str(r.trans_date)]
        for r in records
    ]
    return create_export(FINANCE_HEADERS, rows, "财务记录.xlsx", FINANCE_COLUMN_WIDTHS)


# ========== Finance Reports ==========

def _parse_query_date(value, label):
    if not value:
        return None
    try:
        return _parse_trans_date(value)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=f"{label}格式错误，请使用 YYYY-MM-DD") from exc


def _build_finance_summary_query(db, current_user, year, month, mine_id, fleet_id=None, start_date=None, end_date=None):
    start_day = _parse_query_date(start_date, "开始日期")
    end_day = _parse_query_date(end_date, "结束日期")
    query = db.query(FinanceRecord)
    query = _filter_scope(query, current_user, mine_id, fleet_id)

    if start_day:
        query = query.filter(FinanceRecord.trans_date >= start_day)
    if end_day:
        query = query.filter(FinanceRecord.trans_date <= end_day)

    if not start_day and not end_day and (year or month):
        target_year = year or date.today().year
        query = query.filter(extract('year', FinanceRecord.trans_date) == target_year)
        if month:
            query = query.filter(extract('month', FinanceRecord.trans_date) == month)
        year = target_year

    return query, year, month, start_day, end_day


def _empty_currency_bucket():
    return {"income": 0.0, "expense": 0.0}


def _summary_from_records(records):
    summary = defaultdict(_empty_currency_bucket)
    for item in records:
        currency = item.currency or "USD"
        trans_type = item.trans_type or ""
        if trans_type not in ("income", "expense"):
            continue
        summary[currency][trans_type] += float(item.amount or 0)
    return {currency: values for currency, values in summary.items()}


def _expense_description(record):
    description = (record.description or "").strip()
    if description:
        return description
    if record.trans_type == "expense":
        return "未填写支出说明"
    return "未填写收入说明"


def _aggregate_expenses(records, key_getter, label_key):
    grouped = {}
    for item in records:
        if item.trans_type != "expense":
            continue
        label = key_getter(item) or "未分类"
        currency = item.currency or "USD"
        key = (label, currency)
        if key not in grouped:
            grouped[key] = {label_key: label, "currency": currency, "amount": 0.0, "count": 0}
        grouped[key]["amount"] += float(item.amount or 0)
        grouped[key]["count"] += 1
    return sorted(grouped.values(), key=lambda row: row["amount"], reverse=True)


def _monthly_trend(records):
    grouped = {}
    for item in records:
        if item.trans_type not in ("income", "expense"):
            continue
        month_key = item.trans_date.strftime("%Y-%m")
        currency = item.currency or "USD"
        key = (month_key, currency)
        if key not in grouped:
            grouped[key] = {"month": month_key, "currency": currency, "income": 0.0, "expense": 0.0}
        grouped[key][item.trans_type] += float(item.amount or 0)
    return sorted(grouped.values(), key=lambda row: (row["month"], row["currency"]))


@router.get("/summary/by-currency")
def get_finance_summary_by_currency(
    mine_id: str = None,
    fleet_id: str = None,
    year: int = None,
    month: int = None,
    start_date: str = None,
    end_date: str = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user)
):
    base_query, year, month, start_day, end_day = _build_finance_summary_query(
        db, current_user, year, month, mine_id, fleet_id, start_date, end_date
    )

    results = base_query.with_entities(
        FinanceRecord.currency,
        FinanceRecord.trans_type,
        func.sum(FinanceRecord.amount).label("total")
    ).group_by(FinanceRecord.currency, FinanceRecord.trans_type).all()

    summary = {}
    for r in results:
        currency = r.currency
        if currency not in summary:
            summary[currency] = {"income": 0, "expense": 0}
        summary[currency][r.trans_type] = float(r.total or 0)

    return {
        "year": year,
        "month": month,
        "start_date": str(start_day) if start_day else None,
        "end_date": str(end_day) if end_day else None,
        "currencies": summary
    }


@router.get("/summary/expense-by-category")
def get_expense_by_category(
    mine_id: str = None,
    fleet_id: str = None,
    year: int = None,
    month: int = None,
    start_date: str = None,
    end_date: str = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user)
):
    base_query, year, month, start_day, end_day = _build_finance_summary_query(
        db, current_user, year, month, mine_id, fleet_id, start_date, end_date
    )

    results = base_query.with_entities(
        FinanceRecord.category,
        FinanceRecord.currency,
        func.sum(FinanceRecord.amount).label("total")
    ).filter(
        FinanceRecord.trans_type == "expense"
    ).group_by(FinanceRecord.category, FinanceRecord.currency).all()

    categories = [
        {
            "category": r.category or "未分类",
            "currency": r.currency,
            "amount": float(r.total or 0)
        }
        for r in results
    ]

    by_currency = {}
    for item in categories:
        by_currency.setdefault(item["currency"], []).append({
            "category": item["category"],
            "amount": item["amount"]
        })

    return {
        "year": year,
        "month": month,
        "start_date": str(start_day) if start_day else None,
        "end_date": str(end_day) if end_day else None,
        "categories": categories,
        "by_currency": by_currency
    }


@router.get("/analysis")
def get_finance_analysis(
    mine_id: str = None,
    fleet_id: str = None,
    year: int = None,
    month: int = None,
    start_date: str = None,
    end_date: str = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user)
):
    base_query, year, month, start_day, end_day = _build_finance_summary_query(
        db, current_user, year, month, mine_id, fleet_id, start_date, end_date
    )
    records = base_query.order_by(FinanceRecord.trans_date.asc(), FinanceRecord.created_at.asc()).all()
    income_count = sum(1 for item in records if item.trans_type == "income")
    expense_count = sum(1 for item in records if item.trans_type == "expense")
    return {
        "year": year,
        "month": month,
        "start_date": str(start_day) if start_day else None,
        "end_date": str(end_day) if end_day else None,
        "record_count": len(records),
        "income_count": income_count,
        "expense_count": expense_count,
        "currencies": _summary_from_records(records),
        "monthly_trend": _monthly_trend(records),
        "expense_categories": _aggregate_expenses(records, lambda item: item.category or "未分类", "category"),
        "expense_descriptions": _aggregate_expenses(records, _expense_description, "description")
    }


# ========== Finance CRUD (Dynamic routes must stay last) ==========

@router.get("/{finance_id}", response_model=FinanceOut)
def read_finance(
    finance_id: str,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user)
):
    return _get_and_check(db, FinanceRecord, finance_id, current_user)


@router.put("/{finance_id}", response_model=FinanceOut)
def update_finance(
    finance_id: str,
    finance_update: FinanceUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user)
):
    db_finance = _get_and_check(db, FinanceRecord, finance_id, current_user)
    assert_period_unlocked(db, db_finance.mine_id, "finance", db_finance.trans_date)
    if finance_update.trans_date is not None:
        assert_period_unlocked(db, db_finance.mine_id, "finance", finance_update.trans_date)
    next_trans_type = _enum_value(finance_update.trans_type) if finance_update.trans_type is not None else db_finance.trans_type
    next_category = finance_update.category if finance_update.category is not None else db_finance.category
    normalized_category = _normalize_category(next_trans_type, next_category)

    for field in FINANCE_FIELDS:
        val = getattr(finance_update, field, None)
        if val is not None:
            setattr(db_finance, field, _enum_value(val))
    db_finance.category = normalized_category
    db.commit()
    db.refresh(db_finance)
    return db_finance


@router.delete("/{finance_id}")
def delete_finance(
    finance_id: str,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user)
):
    db_finance = _get_and_check(db, FinanceRecord, finance_id, current_user)
    assert_period_unlocked(db, db_finance.mine_id, "finance", db_finance.trans_date)
    archive_record(db, db_finance, "finance", current_user)
    db.delete(db_finance)
    db.commit()
    return {"message": "财务记录已删除"}
