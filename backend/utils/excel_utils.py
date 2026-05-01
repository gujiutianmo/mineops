"""统一的Excel导入导出工具"""
import io
import re
from typing import List, Callable, Optional
from urllib.parse import quote
from fastapi import UploadFile, HTTPException
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center")


def create_template(headers: List[str], filename: str, column_widths: List[int] = None):
    """生成导入模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Import Template"

    # 应用表头样式（与导出一致）
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT

    # 添加示例行
    for col in range(1, len(headers) + 1):
        example_cell = ws.cell(row=2, column=col, value="在此填写数据")
        example_cell.font = Font(italic=True, color="999999")
        example_cell.alignment = Alignment(horizontal="left")

    if column_widths:
        for col_idx, width in enumerate(column_widths, 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    encoded_filename = quote(filename)
    # RFC 6266: provide fallback ASCII filename for legacy clients
    ascii_filename = re.sub(r'[^\x00-\x7F]+', '', filename)
    if not ascii_filename or ascii_filename.startswith('_'):
        ascii_filename = 'download.xlsx'
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=\"{ascii_filename}\"; filename*=UTF-8''{encoded_filename}",
            "Access-Control-Expose-Headers": "Content-Disposition",
            "Cache-Control": "no-cache"
        }
    )


def create_export(headers: List[str], rows: List[List], filename: str, column_widths: List[int] = None):
    """导出数据为Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Export"

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    if column_widths:
        for col_idx, width in enumerate(column_widths, 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    encoded_filename = quote(filename)
    ascii_filename = re.sub(r'[^\x00-\x7F]+', '', filename)
    if not ascii_filename or ascii_filename.startswith('_'):
        ascii_filename = 'export.xlsx'
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=\"{ascii_filename}\"; filename*=UTF-8''{encoded_filename}",
            "Access-Control-Expose-Headers": "Content-Disposition",
            "Cache-Control": "no-cache"
        }
    )


async def parse_import_file(file: UploadFile, expected_headers: List[str]):
    """解析上传的Excel文件，验证表头并返回工作表"""
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Only .xlsx files are accepted")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    actual_headers = [ws.cell(row=1, column=c).value for c in range(1, len(expected_headers) + 1)]
    if actual_headers != expected_headers:
        raise HTTPException(
            status_code=400,
            detail=f"Header mismatch: expected {expected_headers}, got {actual_headers}"
        )

    return ws
